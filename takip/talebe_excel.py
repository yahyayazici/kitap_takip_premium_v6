from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from django.db import transaction

from .models import DiniDersSeviyesi, EtutHocasi, SinifSube, Talebe
from .tc_util import tc_normalize as _tc_normalize
from .telefon_util import telefon_formatla
from .turkiye_il_ilce import memleket_gecerli, tum_ilceler, turkiye_illeri
from .veli_hesap_util import veli_panel_ensure as _veli_panel_ensure
from .wave0_models import VeliKisi

EXCEL_BASLIKLAR = [
    "Talebe No",
    "Kimlik Adı",
    "Kimlik Soyadı",
    "TC Kimlik",
    "Cinsiyet",
    "Doğum Tarihi",
    "Baba Adı",
    "Anne Adı",
    "Doğum Yeri",
    "Memleket İl",
    "Memleket İlçe",
    "Telefon",
    "Okul Seviyesi",
    "Okul Sınıf",
    "Okul Şube",
    "Etüt Mesulü",
    "Dini Ders Seviyesi",
    "Dini Ders Hocası",
    "Aile Durumu",
    "Anne Ad Soyad",
    "Anne Telefon",
    "Baba Ad Soyad",
    "Baba Telefon",
    "Ev Adresi",
    "Aktif",
]

ORNEK_SATIR = [
    "1",
    "Ahmet",
    "Yılmaz",
    "12345678901",
    "Erkek",
    date(2010, 5, 15),
    "Mehmet",
    "Ayşe",
    "İstanbul",
    "Trabzon",
    "Merkez",
    "0532 123 45 67",
    "Ortaokul 5",
    "5",
    "A",
    "Yahya Yazıcı",
    "Ortaokul Seviye 1",
    "Yahya Yazıcı",
    "Anne – baba beraber",
    "Ayşe Yılmaz",
    "0532 111 22 33",
    "Mehmet Yılmaz",
    "0532 444 55 66",
    "Örnek Mah. Örnek Cad. No:1 İstanbul",
    "Evet",
]


@dataclass
class TalebeExcelSonuc:
    eklenen: int = 0
    guncellenen: int = 0
    veli_hesap: int = 0
    atlanan: int = 0
    hatalar: list[str] = field(default_factory=list)
    bilgi: list[str] = field(default_factory=list)


def _hucre_degeri(deger) -> str:
    if deger is None:
        return ""
    if isinstance(deger, datetime):
        return deger.date().isoformat()
    if isinstance(deger, date):
        return deger.isoformat()
    if isinstance(deger, float) and deger.is_integer():
        return str(int(deger))
    return str(deger).strip()


def _aktif_mi(deger: str) -> bool:
    if not deger:
        return True

    normalized = deger.strip().lower()
    if normalized in {"evet", "e", "1", "true", "aktif", "yes", "y"}:
        return True
    if normalized in {"hayır", "hayir", "h", "0", "false", "pasif", "no", "n"}:
        return False
    raise ValueError(f"Geçersiz aktif değeri: {deger}")


def _baslik_eslestir(satir: list[str]) -> dict[str, int]:
    eslesme = {}
    for index, hucre in enumerate(satir):
        anahtar = _hucre_degeri(hucre).lower()
        if anahtar in {
            "ad soyad",
            "ad_soyad",
            "adsoyad",
            "isim",
            "kullanılan ad soyad",
            "kullanilan ad soyad",
        }:
            eslesme["ad_soyad"] = index
        elif anahtar in {"kimlik adı", "kimlik adi", "kimlik_adi"}:
            eslesme["kimlik_adi"] = index
        elif anahtar in {"kimlik soyadı", "kimlik soyadi", "kimlik_soyadi"}:
            eslesme["kimlik_soyadi"] = index
        elif anahtar in {"sınıf", "sinif", "class", "okul sınıf", "okul sinif"}:
            eslesme["sinif"] = index
        elif anahtar in {"şube", "sube", "okul şube", "okul sube"}:
            eslesme["sube"] = index
        elif anahtar in {"talebe tc", "talebe_tc", "tc", "tc kimlik", "tc_kimlik"}:
            eslesme["talebe_tc"] = index
        elif anahtar in {"talebe telefon", "talebe_telefon", "telefon"}:
            eslesme["talebe_telefon"] = index
        elif anahtar in {"anne ad soyad", "anne_ad_soyad", "anne"}:
            eslesme["anne_ad"] = index
        elif anahtar in {"anne telefon", "anne_telefon"}:
            eslesme["anne_telefon"] = index
        elif anahtar in {"baba ad soyad", "baba_ad_soyad", "baba"}:
            eslesme["baba_ad"] = index
        elif anahtar in {"baba telefon", "baba_telefon"}:
            eslesme["baba_telefon"] = index
        elif anahtar in {
            "etüt mesulü",
            "etut mesulu",
            "etüt hocası",
            "etut hocasi",
            "etut_hocasi",
            "hoca",
        }:
            eslesme["etut_hocasi"] = index
        elif anahtar in {
            "dini ders seviyesi",
            "dini_ders_seviyesi",
            "dini seviye",
            "dini seviyesi",
        }:
            eslesme["dini_ders_seviyesi"] = index
        elif anahtar in {
            "dini ders hocası",
            "dini ders hocasi",
            "dini_ders_hocasi",
            "dini hoca",
        }:
            eslesme["dini_ders_hocasi"] = index
        elif anahtar in {"talebe no", "talebe_no", "numara", "no"}:
            eslesme["talebe_no"] = index
        elif anahtar in {"aktif", "durum"}:
            eslesme["aktif"] = index
        elif anahtar in {"cinsiyet"}:
            eslesme["cinsiyet"] = index
        elif anahtar in {"doğum tarihi", "dogum tarihi", "dogum_tarihi"}:
            eslesme["dogum_tarihi"] = index
        elif anahtar in {"baba adı", "baba adi", "baba_adi"}:
            eslesme["baba_adi"] = index
        elif anahtar in {"anne adı", "anne adi", "anne_adi"}:
            eslesme["anne_adi"] = index
        elif anahtar in {"doğum yeri", "dogum yeri", "dogum_yeri"}:
            eslesme["dogum_yeri"] = index
        elif anahtar in {"memleket il", "memleket ili", "memleket"}:
            eslesme["memleket"] = index
        elif anahtar in {"memleket ilçe", "memleket ilce", "memleket_ilce"}:
            eslesme["memleket_ilce"] = index
        elif anahtar in {"okul seviyesi", "okul_seviyesi"}:
            eslesme["okul_seviyesi"] = index
        elif anahtar in {"aile durumu", "aile_durumu"}:
            eslesme["aile_durumu"] = index
        elif anahtar in {"ev adresi", "veli ev adresi", "adres"}:
            eslesme["ev_adresi"] = index
    return eslesme


def _satir_degeri(satir: list, index: int | None) -> str:
    if index is None or index >= len(satir):
        return ""
    return _hucre_degeri(satir[index])


def _cinsiyet_eslestir(deger: str) -> str:
    normalized = (deger or "").strip().lower()
    if normalized in {"erkek", "e", "male", "m"}:
        return Talebe.Cinsiyet.ERKEK
    if normalized in {"kadın", "kadin", "k", "female", "f"}:
        return Talebe.Cinsiyet.KADIN
    return ""


def _aile_durumu_eslestir(deger: str) -> str:
    normalized = (deger or "").strip().lower()
    for choice in Talebe.AileDurumu:
        if normalized == choice.label.lower():
            return choice.value
    eski = {
        "anne – baba beraber": Talebe.AileDurumu.BERABER,
        "anne - baba beraber": Talebe.AileDurumu.BERABER,
        "anne – baba ayrı": Talebe.AileDurumu.AYRI,
        "anne - baba ayrı": Talebe.AileDurumu.AYRI,
        "anne – baba ayrı – baba üvey": Talebe.AileDurumu.AYRI_BABA_UVEY,
        "anne – baba ayrı – anne üvey": Talebe.AileDurumu.AYRI_ANNE_UVEY,
        "anne vefat": Talebe.AileDurumu.ANNE_VEFAT,
        "anne vefat – anne üvey": Talebe.AileDurumu.ANNE_VEFAT_ANNE_UVEY,
    }
    return eski.get(normalized, "")


def _dogum_tarihi_eslestir(deger) -> date | None:
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    metin = str(deger or "").strip()
    if not metin:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(metin, fmt).date()
        except ValueError:
            continue
    return None


def _ad_soyad_olustur(satir: list, basliklar: dict[str, int]) -> str:
    ad_soyad = _satir_degeri(satir, basliklar.get("ad_soyad"))
    if ad_soyad:
        return ad_soyad
    kimlik_adi = _satir_degeri(satir, basliklar.get("kimlik_adi"))
    kimlik_soyadi = _satir_degeri(satir, basliklar.get("kimlik_soyadi"))
    return f"{kimlik_adi} {kimlik_soyadi}".strip()


def _telefon_al(satir: list, basliklar: dict[str, int], anahtar: str) -> str:
    return telefon_formatla(_satir_degeri(satir, basliklar.get(anahtar)))


def _talebe_profil_satirdan(talebe: Talebe, satir: list, basliklar: dict[str, int]) -> bool:
    degisti = False

    def _guncelle(alan: str, deger: str) -> None:
        nonlocal degisti
        if deger and getattr(talebe, alan) != deger:
            setattr(talebe, alan, deger)
            degisti = True

    kimlik_adi = _satir_degeri(satir, basliklar.get("kimlik_adi"))
    kimlik_soyadi = _satir_degeri(satir, basliklar.get("kimlik_soyadi"))
    ad_soyad = _ad_soyad_olustur(satir, basliklar)

    _guncelle("kimlik_adi", kimlik_adi)
    _guncelle("kimlik_soyadi", kimlik_soyadi)
    if ad_soyad:
        _guncelle("ad_soyad", ad_soyad)

    cinsiyet = _cinsiyet_eslestir(_satir_degeri(satir, basliklar.get("cinsiyet")))
    if cinsiyet:
        _guncelle("cinsiyet", cinsiyet)

    dogum_raw = None
    idx = basliklar.get("dogum_tarihi")
    if idx is not None and idx < len(satir):
        dogum_raw = satir[idx]
    dogum = _dogum_tarihi_eslestir(dogum_raw)
    if dogum and talebe.dogum_tarihi != dogum:
        talebe.dogum_tarihi = dogum
        degisti = True

    for alan in ("baba_adi", "anne_adi", "dogum_yeri", "okul_seviyesi", "ev_adresi"):
        _guncelle(alan, _satir_degeri(satir, basliklar.get(alan)))

    memleket = _satir_degeri(satir, basliklar.get("memleket"))
    memleket_ilce = _satir_degeri(satir, basliklar.get("memleket_ilce"))
    if memleket:
        if memleket_ilce and not memleket_gecerli(memleket, memleket_ilce):
            pass
        else:
            _guncelle("memleket", memleket)
            if memleket_ilce:
                _guncelle("memleket_ilce", memleket_ilce)

    aile = _aile_durumu_eslestir(_satir_degeri(satir, basliklar.get("aile_durumu")))
    if aile:
        _guncelle("aile_durumu", aile)

    return degisti


def _excel_satirlari(dosya) -> list[list]:
    from openpyxl import load_workbook

    workbook = load_workbook(dosya, read_only=True, data_only=True)
    sayfa = workbook.active
    satirlar = []

    for satir in sayfa.iter_rows(values_only=True):
        if not satir:
            continue
        degerler = list(satir)
        if any(_hucre_degeri(h) for h in degerler):
            satirlar.append(degerler)

    workbook.close()
    return satirlar


def _xlsx_kaydet(
    satirlar: list[list],
    *,
    sayfa_adi: str = "Talebeler",
    dogrulama: bool = True,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    sayfa = workbook.active
    sayfa.title = sayfa_adi

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill("solid", fgColor="1A4FA8")

    for row_idx, satir in enumerate(satirlar, start=1):
        for col_idx, deger in enumerate(satir, start=1):
            hucre = sayfa.cell(row=row_idx, column=col_idx, value=deger)
            if row_idx == 1:
                hucre.font = baslik_font
                hucre.fill = baslik_fill

    genislikler = [
        10, 14, 14, 14, 10, 14, 14, 14, 14, 14, 14, 16, 14, 10, 10, 20, 20, 20, 22,
        18, 16, 18, 16, 28, 8,
    ]
    for col, genislik in enumerate(genislikler, start=1):
        sayfa.column_dimensions[get_column_letter(col)].width = genislik

    baslik_map = {
        str(baslik).strip().lower(): idx + 1
        for idx, baslik in enumerate(satirlar[0])
    } if satirlar else {}
    kol_dogum = baslik_map.get("doğum tarihi") or baslik_map.get("dogum tarihi")
    if kol_dogum:
        for row_idx in range(2, max(len(satirlar) + 200, 500) + 1):
            hucre = sayfa.cell(row=row_idx, column=kol_dogum)
            hucre.number_format = "DD.MM.YYYY"

    if dogrulama and satirlar:
        liste = workbook.create_sheet("_Listeler")
        liste.sheet_state = "hidden"

        siniflar = sorted(
            {
                ss.sinif.strip()
                for ss in SinifSube.objects.filter(aktif=True)
                if ss.sinif.strip()
            }
        )
        subeler = sorted(
            {
                ss.sube.strip()
                for ss in SinifSube.objects.filter(aktif=True)
                if ss.sube.strip()
            }
        )
        seviyeler = list(
            DiniDersSeviyesi.objects.filter(aktif=True)
            .order_by("sira", "ad")
            .values_list("ad", flat=True)
        )
        hocalar = list(
            EtutHocasi.objects.filter(aktif=True)
            .order_by("ad_soyad")
            .values_list("ad_soyad", flat=True)
        )
        iller = turkiye_illeri()
        ilceler = tum_ilceler()
        cinsiyetler = ["Erkek", "Kadın"]
        aile_secenekleri = [c.label for c in Talebe.AileDurumu]
        aktif_secenekleri = ["Evet", "Hayır"]

        listeler = [
            siniflar,
            subeler,
            seviyeler,
            hocalar,
            cinsiyetler,
            aile_secenekleri,
            aktif_secenekleri,
            iller,
            ilceler,
        ]
        for kol_idx, degerler in enumerate(listeler):
            for satir_idx, deger in enumerate(degerler, start=1):
                liste.cell(row=satir_idx, column=kol_idx + 1, value=deger)

        def _liste_formulu(kol_harf: str, uzunluk: int) -> str:
            if uzunluk <= 0:
                return '""'
            return f"=_Listeler!${kol_harf}$1:${kol_harf}${uzunluk}"

        def _dogrulama_ekle(formul: str, hucre_araligi: str) -> None:
            if formul == '""':
                return
            dv = DataValidation(
                type="list",
                formula1=formul,
                allow_blank=True,
                showDropDown=False,
            )
            dv.errorStyle = "warning"
            dv.showErrorMessage = False
            sayfa.add_data_validation(dv)
            dv.add(hucre_araligi)

        son_satir = max(len(satirlar) + 200, 500)

        def _kolon(baslik: str) -> str | None:
            idx = baslik_map.get(baslik.lower())
            if not idx:
                return None
            return get_column_letter(idx)

        eslesmeler = [
            ("Okul Sınıf", "A", len(siniflar)),
            ("Okul Şube", "B", len(subeler)),
            ("Dini Ders Seviyesi", "C", len(seviyeler)),
            ("Etüt Mesulü", "D", len(hocalar)),
            ("Dini Ders Hocası", "D", len(hocalar)),
            ("Cinsiyet", "E", len(cinsiyetler)),
            ("Aile Durumu", "F", len(aile_secenekleri)),
            ("Aktif", "G", len(aktif_secenekleri)),
            ("Memleket İl", "H", len(iller)),
            ("Memleket İlçe", "I", len(ilceler)),
        ]
        for baslik, harf, uzunluk in eslesmeler:
            kol = _kolon(baslik)
            if kol:
                _dogrulama_ekle(
                    _liste_formulu(harf, uzunluk),
                    f"{kol}2:{kol}{son_satir}",
                )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sablon_xlsx_olustur() -> bytes:
    return _xlsx_kaydet([EXCEL_BASLIKLAR, ORNEK_SATIR])


def mevcut_talebeler_xlsx_olustur(*, talebe_qs=None) -> bytes:
    if talebe_qs is None:
        talebe_qs = (
            Talebe.objects.filter(aktif=True)
            .select_related(
                "sinif_sube",
                "etut_hocasi",
                "dini_ders_hocasi",
                "dini_ders_seviyesi",
            )
            .prefetch_related("veli_kisileri")
            .order_by("sinif", "sube", "ad_soyad")
        )

    satirlar = [EXCEL_BASLIKLAR]
    for talebe in talebe_qs:
        anne = baba = None
        for veli in talebe.veli_kisileri.all():
            if veli.yakinlik == VeliKisi.Yakinlik.ANNE:
                anne = veli
            elif veli.yakinlik == VeliKisi.Yakinlik.BABA:
                baba = veli

        sinif = talebe.sinif_sube.sinif if talebe.sinif_sube_id else talebe.sinif
        sube = talebe.sinif_sube.sube if talebe.sinif_sube_id else talebe.sube
        aile_etiket = talebe.get_aile_durumu_display() if talebe.aile_durumu else ""
        cinsiyet_etiket = talebe.get_cinsiyet_display() if talebe.cinsiyet else ""
        dogum = talebe.dogum_tarihi if talebe.dogum_tarihi else ""
        satirlar.append(
            [
                talebe.talebe_no or "",
                talebe.kimlik_adi or "",
                talebe.kimlik_soyadi or "",
                talebe.tc_kimlik or "",
                cinsiyet_etiket,
                dogum,
                talebe.baba_adi or "",
                talebe.anne_adi or "",
                talebe.dogum_yeri or "",
                talebe.memleket or "",
                talebe.memleket_ilce or "",
                telefon_formatla(talebe.telefon) if talebe.telefon else "",
                talebe.okul_seviyesi or "",
                sinif,
                sube,
                talebe.etut_hocasi.ad_soyad if talebe.etut_hocasi_id else "",
                str(talebe.dini_ders_seviyesi) if talebe.dini_ders_seviyesi_id else "",
                talebe.dini_ders_hocasi.ad_soyad if talebe.dini_ders_hocasi_id else "",
                aile_etiket,
                anne.ad_soyad if anne else "",
                telefon_formatla(anne.telefon) if anne and anne.telefon else "",
                baba.ad_soyad if baba else "",
                telefon_formatla(baba.telefon) if baba and baba.telefon else "",
                talebe.ev_adresi or "",
                "Evet" if talebe.aktif else "Hayır",
            ]
        )

    return _xlsx_kaydet(satirlar)


def _veli_kisi_guncelle(
    talebe: Talebe,
    yakinlik: str,
    ad_soyad: str,
    telefon: str,
) -> None:
    if not ad_soyad:
        return

    telefon = telefon_formatla(telefon) if telefon else ""
    veli = VeliKisi.objects.filter(talebe=talebe, yakinlik=yakinlik).first()
    if veli:
        veli.ad_soyad = ad_soyad
        if telefon:
            veli.telefon = telefon
        veli.save(update_fields=["ad_soyad", "telefon"])
        return

    VeliKisi.objects.create(
        talebe=talebe,
        yakinlik=yakinlik,
        ad_soyad=ad_soyad,
        telefon=telefon,
        birincil=yakinlik == VeliKisi.Yakinlik.ANNE,
    )


def _talebe_bul(
    *,
    talebe_no: str,
    ad_soyad: str,
    sinif: str,
    sube: str,
    tc: str = "",
) -> Talebe | None:
    if talebe_no:
        talebe = Talebe.objects.filter(talebe_no=talebe_no).first()
        if talebe:
            return talebe

    if tc and len(tc) == 11:
        talebe = Talebe.objects.filter(tc_kimlik=tc).first()
        if talebe:
            return talebe

    if ad_soyad and sinif and sube:
        return (
            Talebe.objects.filter(
                ad_soyad__iexact=ad_soyad,
                sinif__iexact=sinif,
                sube__iexact=sube,
            )
            .select_related("sinif_sube", "etut_hocasi")
            .first()
        )
    return None


@transaction.atomic
def talebe_excel_ice_aktar(dosya) -> TalebeExcelSonuc:
    sonuc = TalebeExcelSonuc()

    try:
        satirlar = _excel_satirlari(dosya)
    except Exception as exc:
        sonuc.hatalar.append(f"Dosya okunamadı: {exc}")
        return sonuc

    if not satirlar:
        sonuc.hatalar.append("Excel dosyası boş görünüyor.")
        return sonuc

    basliklar = _baslik_eslestir(
        [_hucre_degeri(h) for h in satirlar[0]]
    )
    if (
        "ad_soyad" not in basliklar
        and not ("kimlik_adi" in basliklar and "kimlik_soyadi" in basliklar)
    ):
        sonuc.hatalar.append(
            "Başlık satırında «Kimlik Adı» ve «Kimlik Soyadı» (veya Ad Soyad) olmalı."
        )
        return sonuc

    hoca_haritasi = {
        hoca.ad_soyad.strip().lower(): hoca
        for hoca in EtutHocasi.objects.filter(aktif=True)
    }
    seviye_haritasi = {
        seviye.ad.strip().lower(): seviye
        for seviye in DiniDersSeviyesi.objects.filter(aktif=True).prefetch_related(
            "hocalar"
        )
    }
    sinif_haritasi = {
        (grup.sinif.strip().lower(), grup.sube.strip().lower()): grup
        for grup in SinifSube.objects.filter(aktif=True)
    }

    siradaki_no = Talebe._yeni_talebe_no()

    def _sonraki_no() -> str:
        nonlocal siradaki_no
        mevcut = siradaki_no
        siradaki_no = str(int(siradaki_no) + 1)
        return mevcut

    for satir_no, satir in enumerate(satirlar[1:], start=2):
        ad_soyad = _ad_soyad_olustur(satir, basliklar)
        talebe_no = _satir_degeri(satir, basliklar.get("talebe_no"))
        sinif = _satir_degeri(satir, basliklar.get("sinif"))
        sube = _satir_degeri(satir, basliklar.get("sube"))
        talebe_tc = _tc_normalize(_satir_degeri(satir, basliklar.get("talebe_tc")))
        talebe_telefon = _telefon_al(satir, basliklar, "talebe_telefon")
        anne_ad = _satir_degeri(satir, basliklar.get("anne_ad"))
        anne_tel = _telefon_al(satir, basliklar, "anne_telefon")
        baba_ad = _satir_degeri(satir, basliklar.get("baba_ad"))
        baba_tel = _telefon_al(satir, basliklar, "baba_telefon")
        hoca_adi = _satir_degeri(satir, basliklar.get("etut_hocasi"))
        dini_seviye_adi = _satir_degeri(satir, basliklar.get("dini_ders_seviyesi"))
        dini_hoca_adi = _satir_degeri(satir, basliklar.get("dini_ders_hocasi"))
        aktif_raw = _satir_degeri(satir, basliklar.get("aktif"))
        memleket = _satir_degeri(satir, basliklar.get("memleket"))
        memleket_ilce = _satir_degeri(satir, basliklar.get("memleket_ilce"))

        if memleket and memleket_ilce and not memleket_gecerli(memleket, memleket_ilce):
            sonuc.bilgi.append(
                f"Satır {satir_no}: Memleket il/ilçe uyuşmuyor "
                f"({memleket}/{memleket_ilce})."
            )

        sinif_sube = None
        if sinif and sube:
            sinif_sube = sinif_haritasi.get((sinif.lower(), sube.lower()))

        if not hoca_adi and sinif_sube:
            zimmet_hocalar = list(
                sinif_sube.etut_hocalari.filter(aktif=True).order_by("ad_soyad")
            )
            if len(zimmet_hocalar) == 1:
                hoca_adi = zimmet_hocalar[0].ad_soyad

        dini_seviye = (
            seviye_haritasi.get(dini_seviye_adi.lower()) if dini_seviye_adi else None
        )
        dini_hoca = hoca_haritasi.get(dini_hoca_adi.lower()) if dini_hoca_adi else None

        if not ad_soyad and not talebe_no:
            continue

        mevcut = _talebe_bul(
            talebe_no=talebe_no,
            ad_soyad=ad_soyad,
            sinif=sinif,
            sube=sube,
            tc=talebe_tc,
        )

        if mevcut:
            degisti = _talebe_profil_satirdan(mevcut, satir, basliklar)
            islem_yapildi = False

            if talebe_tc and len(talebe_tc) == 11 and mevcut.tc_kimlik != talebe_tc:
                mevcut.tc_kimlik = talebe_tc
                degisti = True
            elif talebe_tc and len(talebe_tc) != 11:
                sonuc.bilgi.append(
                    f"Satır {satir_no}: TC geçersiz ({talebe_tc}), atlandı."
                )

            if talebe_telefon and mevcut.telefon != talebe_telefon:
                mevcut.telefon = talebe_telefon
                degisti = True

            if aktif_raw:
                try:
                    aktif = _aktif_mi(aktif_raw)
                    if mevcut.aktif != aktif:
                        mevcut.aktif = aktif
                        degisti = True
                except ValueError as exc:
                    sonuc.bilgi.append(f"Satır {satir_no}: {exc}")

            if dini_seviye_adi and not dini_seviye:
                sonuc.bilgi.append(
                    f"Satır {satir_no}: '{dini_seviye_adi}' dini ders seviyesi bulunamadı."
                )
            elif dini_seviye and mevcut.dini_ders_seviyesi_id != dini_seviye.pk:
                mevcut.dini_ders_seviyesi = dini_seviye
                degisti = True

            if dini_hoca_adi and not dini_hoca:
                sonuc.bilgi.append(
                    f"Satır {satir_no}: '{dini_hoca_adi}' dini ders hocası bulunamadı."
                )
            elif dini_hoca:
                if dini_seviye and not dini_seviye.hocalar.filter(pk=dini_hoca.pk).exists():
                    sonuc.bilgi.append(
                        f"Satır {satir_no}: {dini_hoca.ad_soyad} «{dini_seviye.ad}» "
                        "seviyesinden sorumlu değil."
                    )
                elif mevcut.dini_ders_hocasi_id != dini_hoca.pk:
                    mevcut.dini_ders_hocasi = dini_hoca
                    degisti = True

            if hoca_adi:
                etut_hocasi = hoca_haritasi.get(hoca_adi.lower())
                if not etut_hocasi:
                    sonuc.bilgi.append(
                        f"Satır {satir_no}: '{hoca_adi}' etüt hocası bulunamadı."
                    )
                elif mevcut.etut_hocasi_id != etut_hocasi.pk:
                    mevcut.etut_hocasi = etut_hocasi
                    degisti = True

            if degisti:
                mevcut.save()
                islem_yapildi = True

            if anne_ad or anne_tel:
                _veli_kisi_guncelle(mevcut, VeliKisi.Yakinlik.ANNE, anne_ad, anne_tel)
                islem_yapildi = True
            if baba_ad or baba_tel:
                _veli_kisi_guncelle(mevcut, VeliKisi.Yakinlik.BABA, baba_ad, baba_tel)
                islem_yapildi = True

            if talebe_tc and len(talebe_tc) == 11:
                veli_ad = anne_ad or baba_ad or ""
                veli_tel = anne_tel or baba_tel or talebe_telefon
                if _veli_panel_ensure(mevcut, talebe_tc, veli_ad, veli_tel):
                    sonuc.veli_hesap += 1
                    islem_yapildi = True
                else:
                    sonuc.bilgi.append(
                        f"Satır {satir_no}: {talebe_tc} kullanıcı adı başka hesapta, "
                        "veli paneli atlandı."
                    )

            if islem_yapildi:
                sonuc.guncellenen += 1
            continue

        if not ad_soyad:
            sonuc.bilgi.append(
                f"Satır {satir_no}: Talebe bulunamadı, kimlik ad/soyad boş — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if not sinif or not sube or not hoca_adi:
            sonuc.bilgi.append(
                f"Satır {satir_no}: Yeni kayıt için sınıf, şube ve etüt mesulü gerekli — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if not sinif_sube:
            sinif_sube = sinif_haritasi.get((sinif.lower(), sube.lower()))
        if not sinif_sube:
            sonuc.bilgi.append(
                f"Satır {satir_no}: {sinif}/{sube} sınıfı bulunamadı — atlandı."
            )
            sonuc.atlanan += 1
            continue

        etut_hocasi = hoca_haritasi.get(hoca_adi.lower())
        if not etut_hocasi:
            sonuc.bilgi.append(
                f"Satır {satir_no}: '{hoca_adi}' etüt hocası bulunamadı — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if not etut_hocasi.sorumlu_sinif_subeler.filter(pk=sinif_sube.pk).exists():
            sonuc.bilgi.append(
                f"Satır {satir_no}: {etut_hocasi.ad_soyad} bu sınıftan sorumlu değil — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if dini_seviye_adi and not dini_seviye:
            sonuc.bilgi.append(
                f"Satır {satir_no}: '{dini_seviye_adi}' dini ders seviyesi bulunamadı — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if dini_seviye and not dini_hoca_adi:
            sonuc.bilgi.append(
                f"Satır {satir_no}: Dini ders seviyesi girildi; dini ders hocası seçin — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if dini_hoca_adi and not dini_hoca:
            sonuc.bilgi.append(
                f"Satır {satir_no}: '{dini_hoca_adi}' dini ders hocası bulunamadı — atlandı."
            )
            sonuc.atlanan += 1
            continue

        if dini_seviye and dini_hoca:
            if not dini_seviye.hocalar.filter(pk=dini_hoca.pk).exists():
                sonuc.bilgi.append(
                    f"Satır {satir_no}: {dini_hoca.ad_soyad} «{dini_seviye.ad}» "
                    "seviyesinden sorumlu değil — atlandı."
                )
                sonuc.atlanan += 1
                continue

        try:
            aktif = _aktif_mi(aktif_raw)
        except ValueError as exc:
            sonuc.bilgi.append(f"Satır {satir_no}: {exc}")
            aktif = True

        if talebe_no and Talebe.objects.filter(talebe_no=talebe_no).exists():
            sonuc.atlanan += 1
            sonuc.bilgi.append(
                f"Satır {satir_no}: {talebe_no} numarası kullanılıyor — atlandı."
            )
            continue

        if Talebe.objects.filter(
            ad_soyad__iexact=ad_soyad,
            sinif_sube=sinif_sube,
        ).exists():
            sonuc.atlanan += 1
            sonuc.bilgi.append(
                f"Satır {satir_no}: {ad_soyad} zaten kayıtlı — atlandı."
            )
            continue

        if not talebe_no:
            talebe_no = _sonraki_no()

        atanan_dini_hoca = dini_hoca or etut_hocasi
        kimlik_adi = _satir_degeri(satir, basliklar.get("kimlik_adi"))
        kimlik_soyadi = _satir_degeri(satir, basliklar.get("kimlik_soyadi"))

        talebe = Talebe(
            ad_soyad=ad_soyad,
            kimlik_adi=kimlik_adi,
            kimlik_soyadi=kimlik_soyadi,
            talebe_no=talebe_no,
            sinif_sube=sinif_sube,
            etut_hocasi=etut_hocasi,
            dini_ders_hocasi=atanan_dini_hoca,
            dini_ders_seviyesi=dini_seviye,
            aktif=aktif,
            telefon=talebe_telefon,
            tc_kimlik=talebe_tc if len(talebe_tc) == 11 else "",
        )
        talebe.save()
        _talebe_profil_satirdan(talebe, satir, basliklar)
        talebe.save()
        sonuc.eklenen += 1

        _veli_kisi_guncelle(talebe, VeliKisi.Yakinlik.ANNE, anne_ad, anne_tel)
        _veli_kisi_guncelle(talebe, VeliKisi.Yakinlik.BABA, baba_ad, baba_tel)

        if talebe_tc and len(talebe_tc) == 11:
            veli_ad = anne_ad or baba_ad or ""
            veli_tel = anne_tel or baba_tel or talebe_telefon
            if _veli_panel_ensure(talebe, talebe_tc, veli_ad, veli_tel):
                sonuc.veli_hesap += 1

    return sonuc
