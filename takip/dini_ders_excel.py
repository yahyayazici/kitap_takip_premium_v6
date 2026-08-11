"""Dini ders takip alanı / konu listesi Excel şablon ve içe aktarma."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from takip.models import DiniDersKonu, DiniDersSeviyesi, DiniDersTakipAlani

ALAN_BASLIKLAR = ("Alan Adı", "Sıra", "Aktif")
KONU_BASLIKLAR = ("Alan", "Konu", "Sıra", "Aktif")
KONU_TEK_SAYFA_BASLIKLAR = ("Seviye", "Alan", "Konu", "Sıra", "Aktif")

_SEVIYE_SAYFA_ADLARI = ("Seviye 1", "Seviye 2", "Seviye 3", "Seviye 4")


@dataclass
class DiniDersExcelSonuc:
    eklenen: int = 0
    guncellenen: int = 0
    atlanan: int = 0
    hatalar: list[str] = field(default_factory=list)
    bilgi: list[str] = field(default_factory=list)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().casefold().replace("ı", "i")


def _aktif_oku(value: Any, *, varsayilan: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return varsayilan
    metin = str(value).strip().casefold()
    if metin in {"1", "true", "evet", "aktif", "yes", "e", "a"}:
        return True
    if metin in {"0", "false", "hayir", "hayır", "pasif", "no", "h", "p"}:
        return False
    return varsayilan


def _sira_oku(value: Any, *, varsayilan: int = 0) -> int:
    if value is None or str(value).strip() == "":
        return varsayilan
    try:
        return max(0, int(float(str(value).strip().replace(",", "."))))
    except (TypeError, ValueError):
        return varsayilan


def _workbook_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def alan_sablon_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Takip Alanlari"
    ws.append(list(ALAN_BASLIKLAR))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Boş örnek satırlar
    for sira in range(1, 6):
        ws.append(["", sira, "Aktif"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    return _workbook_bytes(wb)


def konu_sablon_xlsx() -> bytes:
    """4 seviye için ayrı sayfalar + tek sayfalık alternatif rehber."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    # İlk sayfa: tüm seviyeler tek listede
    ws_hepsi = wb.active
    ws_hepsi.title = "Tum Seviyeler"
    ws_hepsi.append(list(KONU_TEK_SAYFA_BASLIKLAR))
    for cell in ws_hepsi[1]:
        cell.font = Font(bold=True)
    for seviye_ad in _SEVIYE_SAYFA_ADLARI:
        ws_hepsi.append([seviye_ad, "", "", 1, "Aktif"])
    for col, genislik in zip("ABCDE", (14, 22, 36, 10, 10)):
        ws_hepsi.column_dimensions[col].width = genislik

    for seviye_ad in _SEVIYE_SAYFA_ADLARI:
        ws = wb.create_sheet(seviye_ad)
        ws.append(list(KONU_BASLIKLAR))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for sira in range(1, 8):
            ws.append(["", "", sira, "Aktif"])
        for col, genislik in zip("ABCD", (22, 36, 10, 10)):
            ws.column_dimensions[col].width = genislik

    return _workbook_bytes(wb)


def _satir_dict(headers: list[str], values: list[Any]) -> dict[str, Any]:
    mapping = {}
    for idx, header in enumerate(headers):
        key = _normalize_header(header)
        if not key:
            continue
        mapping[key] = values[idx] if idx < len(values) else None
    return mapping


def _alan_ad_al(row: dict[str, Any]) -> str:
    for key in ("alan adi", "alan adı", "alan", "takip alani", "takip alanı"):
        val = row.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _konu_ad_al(row: dict[str, Any]) -> str:
    for key in ("konu", "konu adi", "konu adı"):
        val = row.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def _seviye_ad_al(row: dict[str, Any]) -> str:
    for key in ("seviye", "dini ders seviyesi", "seviye adi", "seviye adı"):
        val = row.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    return ""


def alan_excel_ice_aktar(dosya) -> DiniDersExcelSonuc:
    from openpyxl import load_workbook

    sonuc = DiniDersExcelSonuc()
    wb = load_workbook(dosya, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sonuc.hatalar.append("Excel boş.")
        return sonuc

    headers = [str(h or "").strip() for h in rows[0]]
    for satir_no, values in enumerate(rows[1:], start=2):
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = _satir_dict(headers, list(values))
        ad = _alan_ad_al(row)
        if not ad:
            sonuc.atlanan += 1
            sonuc.hatalar.append(f"Satır {satir_no}: Alan adı boş.")
            continue
        sira = _sira_oku(row.get("sira"), varsayilan=0)
        aktif = _aktif_oku(row.get("aktif"), varsayilan=True)
        _, created = DiniDersTakipAlani.objects.update_or_create(
            ad=ad,
            defaults={"sira": sira, "aktif": aktif},
        )
        if created:
            sonuc.eklenen += 1
        else:
            sonuc.guncellenen += 1

    sonuc.bilgi.append(
        f"Takip alanları: {sonuc.eklenen} eklendi, {sonuc.guncellenen} güncellendi."
    )
    return sonuc


def _seviye_bul_veya_hata(
    seviye_ad: str,
    *,
    sonuc: DiniDersExcelSonuc,
    satir_no: int | str,
    cache: dict[str, DiniDersSeviyesi],
) -> DiniDersSeviyesi | None:
    if not seviye_ad:
        sonuc.atlanan += 1
        sonuc.hatalar.append(f"Satır {satir_no}: Seviye boş.")
        return None
    key = seviye_ad.casefold()
    if key in cache:
        return cache[key]
    seviye = DiniDersSeviyesi.objects.filter(ad__iexact=seviye_ad).first()
    if not seviye:
        sonuc.atlanan += 1
        sonuc.hatalar.append(
            f"Satır {satir_no}: '{seviye_ad}' seviyesi bulunamadı. "
            "Önce Yönetim → Seviyeler’den tanımlayın (Seviye 1–4)."
        )
        return None
    cache[key] = seviye
    return seviye


def _alan_bul_veya_olustur(alan_ad: str, cache: dict[str, DiniDersTakipAlani]) -> DiniDersTakipAlani:
    key = alan_ad.casefold()
    if key in cache:
        return cache[key]
    alan = DiniDersTakipAlani.objects.filter(ad__iexact=alan_ad).first()
    if alan:
        cache[key] = alan
        return alan
    max_sira = (
        DiniDersTakipAlani.objects.order_by("-sira").values_list("sira", flat=True).first()
        or 0
    )
    alan = DiniDersTakipAlani.objects.create(
        ad=alan_ad,
        sira=max_sira + 1,
        aktif=True,
    )
    cache[key] = alan
    return alan


def _konu_satirini_isle(
    *,
    sonuc: DiniDersExcelSonuc,
    satir_no: int | str,
    seviye_ad: str,
    alan_ad: str,
    konu_ad: str,
    sira: int,
    aktif: bool,
    seviye_cache: dict[str, DiniDersSeviyesi],
    alan_cache: dict[str, DiniDersTakipAlani],
) -> None:
    seviye = _seviye_bul_veya_hata(
        seviye_ad, sonuc=sonuc, satir_no=satir_no, cache=seviye_cache
    )
    if not seviye:
        return
    if not alan_ad:
        sonuc.atlanan += 1
        sonuc.hatalar.append(f"Satır {satir_no}: Alan boş.")
        return
    if not konu_ad:
        sonuc.atlanan += 1
        sonuc.hatalar.append(f"Satır {satir_no}: Konu boş.")
        return

    alan = _alan_bul_veya_olustur(alan_ad, alan_cache)
    _, created = DiniDersKonu.objects.update_or_create(
        alan=alan,
        seviye=seviye,
        ad=konu_ad,
        defaults={"sira": sira, "aktif": aktif},
    )
    if created:
        sonuc.eklenen += 1
    else:
        sonuc.guncellenen += 1


def konu_excel_ice_aktar(dosya) -> DiniDersExcelSonuc:
    from openpyxl import load_workbook

    sonuc = DiniDersExcelSonuc()
    wb = load_workbook(dosya, read_only=True, data_only=True)
    seviye_cache: dict[str, DiniDersSeviyesi] = {}
    alan_cache: dict[str, DiniDersTakipAlani] = {
        a.ad.casefold(): a for a in DiniDersTakipAlani.objects.all()
    }

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        header_keys = {_normalize_header(h) for h in headers if h}
        sayfa_seviye = ws.title.strip() if ws.title.strip() in _SEVIYE_SAYFA_ADLARI else ""
        tek_sayfa = "seviye" in header_keys

        for satir_no, values in enumerate(rows[1:], start=2):
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            row = _satir_dict(headers, list(values))
            alan_ad = _alan_ad_al(row)
            konu_ad = _konu_ad_al(row)
            # Başlık satırı / örnek boş satırlar
            if not konu_ad and not alan_ad:
                continue
            seviye_ad = _seviye_ad_al(row) if tek_sayfa else sayfa_seviye
            if not seviye_ad and sayfa_seviye:
                seviye_ad = sayfa_seviye
            sira = _sira_oku(row.get("sira"), varsayilan=0)
            aktif = _aktif_oku(row.get("aktif"), varsayilan=True)
            _konu_satirini_isle(
                sonuc=sonuc,
                satir_no=f"{ws.title}:{satir_no}",
                seviye_ad=seviye_ad,
                alan_ad=alan_ad,
                konu_ad=konu_ad,
                sira=sira,
                aktif=aktif,
                seviye_cache=seviye_cache,
                alan_cache=alan_cache,
            )

    sonuc.bilgi.append(
        f"Konular: {sonuc.eklenen} eklendi, {sonuc.guncellenen} güncellendi."
    )
    return sonuc
