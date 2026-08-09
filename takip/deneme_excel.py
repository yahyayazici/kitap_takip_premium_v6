"""Deneme Excel parse, eşleştirme ve aktarım.

Desteklenen formatlar:
1) Düz başlık (Ad Soyad + Türkçe Doğru / …)
2) Okyanus genel sonuç raporu (üst satır branş, alt satır Doğru/Yanlış/Boş/Net)
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from typing import Any

from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone

from takip.models import (
    DenemeBransSonucu,
    DenemeEslestirmeAlias,
    DenemeSinavi,
    DenemeSonucu,
    Talebe,
)

BRANS_TANIMLARI: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("turkce", "Türkçe", ("türkçe", "turkce")),
    ("matematik", "Matematik", ("matematik", "mat")),
    ("fen", "Fen Bilimleri", ("fen bilimleri", "fen bilgisi", "fen", "fizik")),
    ("sosyal", "Sosyal Bilgiler", ("sosyal bilgiler", "sosyal", "inkılap", "inkilap")),
    ("din", "Din Kültürü", ("din kültürü", "din kulturu", "din")),
    ("ingilizce", "İngilizce", ("ingilizce", "ingilizce", "ing")),
)

BRANS_KODLARI = [k for k, _, _ in BRANS_TANIMLARI]

# Benzerlik eşiği: altında öneri göstermeyiz
BENZERLIK_ESIGI = 0.62
# Tek güçlü aday varsa UI'da öne çıkarılır
GUCLU_BENZERLIK = 0.78


def normalize_ad(deger: str) -> str:
    if not deger:
        return ""
    # Türkçe harfler ASCII'ye (ı→i); eşleşme ve başlık okuma için
    metin = deger.strip().replace("İ", "i").replace("I", "i").lower()
    metin = metin.translate(
        str.maketrans(
            {
                "ı": "i",
                "ğ": "g",
                "ü": "u",
                "ş": "s",
                "ö": "o",
                "ç": "c",
            }
        )
    )
    metin = unicodedata.normalize("NFKD", metin)
    metin = "".join(c for c in metin if not unicodedata.combining(c))
    metin = re.sub(r"[^a-z0-9\s]", " ", metin)
    metin = re.sub(r"\s+", " ", metin).strip()
    return metin


def _soyad_uyumlu(excel_ad: str, site_ad: str) -> bool:
    """Son kelime (soyad) yeterince benzer değilse öneriyi ele."""
    a = normalize_ad(excel_ad).split()
    b = normalize_ad(site_ad).split()
    if not a or not b:
        return False
    return SequenceMatcher(None, a[-1], b[-1]).ratio() >= 0.72


def _benzerlik_orani(a: str, b: str) -> float:
    na, nb = normalize_ad(a), normalize_ad(b)
    if not na or not nb:
        return 0.0
    temel = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    if ta and tb:
        jaccard = len(ta & tb) / len(ta | tb)
        temel = max(temel, 0.55 * temel + 0.45 * jaccard)
    return temel


def talebe_benzer_adaylar(excel_ad: str, limit: int = 4) -> list[dict]:
    """Excel adıyla benzer aktif talebeleri skorla."""
    adaylar: list[dict] = []
    excel_norm = normalize_ad(excel_ad)
    excel_parca = excel_norm.split()
    excel_adlar = set(excel_parca[:-1]) if len(excel_parca) > 1 else set(excel_parca)

    for t in Talebe.objects.filter(aktif=True).select_related("sinif_sube"):
        site_ad = t.ad_soyad or ""
        oran = _benzerlik_orani(excel_ad, site_ad)
        if oran < BENZERLIK_ESIGI:
            continue
        if oran < GUCLU_BENZERLIK:
            if not _soyad_uyumlu(excel_ad, site_ad):
                continue
            site_parca = normalize_ad(site_ad).split()
            site_adlar = set(site_parca[:-1]) if len(site_parca) > 1 else set(site_parca)
            # Zayıf skorda en az bir ortak ön ad olmalı
            if not (excel_adlar & site_adlar):
                continue
        adaylar.append(
            {
                "id": t.id,
                "ad_soyad": site_ad,
                "sinif": str(t.sinif_sube or t.sinif or ""),
                "oran": int(round(oran * 100)),
            }
        )
    adaylar.sort(key=lambda x: (-x["oran"], x["ad_soyad"]))
    return adaylar[:limit]


def _hucre(deger) -> str:
    if deger is None:
        return ""
    return str(deger).strip()


def _ondalik(deger: str) -> Decimal:
    if not deger:
        return Decimal("0.00")
    metin = deger.strip().replace(" ", "")
    metin = metin.replace(",", ".")
    try:
        return Decimal(metin).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def _tam_sayi(deger: str) -> int:
    if not deger:
        return 0
    metin = deger.strip().replace(",", ".")
    try:
        return max(0, int(float(metin)))
    except (ValueError, TypeError):
        return 0


def _excel_satirlari(dosya) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(dosya, read_only=True, data_only=True)
    sayfa = workbook.active
    satirlar = []
    for satir in sayfa.iter_rows(values_only=True):
        if not satir:
            continue
        degerler = [_hucre(h) for h in satir]
        if any(degerler):
            satirlar.append(degerler)
    workbook.close()
    return satirlar


def _brans_kodu(baslik: str) -> str | None:
    anahtar = normalize_ad(baslik)
    if not anahtar:
        return None
    # "puanlar / sıralamalar" branş değil
    if "puan" in anahtar or "siralama" in anahtar:
        return None
    for kod, _, anahtarlar in BRANS_TANIMLARI:
        if any(a in anahtar for a in anahtarlar):
            return kod
    return None


def _okyanus_format_mi(satirlar: list[list[str]]) -> bool:
    if len(satirlar) < 2:
        return False
    alt = satirlar[1]
    ust = satirlar[0]
    ad_altta = any(
        normalize_ad(c) in {"ad soyad", "adsoyad", "isim", "ogrenci", "öğrenci"}
        for c in alt
    )
    brans_ustte = any(_brans_kodu(c) for c in ust if c)
    return ad_altta and brans_ustte


def _duz_baslik_haritasi(baslik_satir: list[str]) -> dict[str, Any]:
    harita: dict[str, Any] = {"brans": {}}
    for idx, baslik in enumerate(baslik_satir):
        anahtar = normalize_ad(baslik)
        if not anahtar:
            continue
        if anahtar in {"ad soyad", "adsoyad", "isim", "ogrenci", "öğrenci"}:
            harita.setdefault("ad_soyad", idx)
            continue
        if anahtar in {"sinif", "sınıf", "class"}:
            harita.setdefault("sinif", idx)
            continue
        if anahtar in {"puan", "score"}:
            harita.setdefault("puan", idx)
            continue
        if anahtar in {"toplam net", "genel net"}:
            harita.setdefault("toplam", {})["net"] = idx
            continue

        for kod, _, anahtarlar in BRANS_TANIMLARI:
            if any(a in anahtar for a in anahtarlar):
                if "dogru" in anahtar or anahtar.endswith(" d"):
                    harita["brans"].setdefault(kod, {})["dogru"] = idx
                elif "yanlis" in anahtar or "yanlış" in baslik.lower():
                    harita["brans"].setdefault(kod, {})["yanlis"] = idx
                elif "bos" in anahtar or "boş" in baslik.lower():
                    harita["brans"].setdefault(kod, {})["bos"] = idx
                elif "net" in anahtar:
                    harita["brans"].setdefault(kod, {})["net"] = idx
                break

        if "toplam" in anahtar or "genel" in anahtar:
            if "dogru" in anahtar:
                harita.setdefault("toplam", {})["dogru"] = idx
            elif "yanlis" in anahtar or "yanlış" in baslik.lower():
                harita.setdefault("toplam", {})["yanlis"] = idx
            elif "bos" in anahtar or "boş" in baslik.lower():
                harita.setdefault("toplam", {})["bos"] = idx
            elif "net" in anahtar:
                harita.setdefault("toplam", {})["net"] = idx

    return harita


def _okyanus_baslik_haritasi(ust: list[str], alt: list[str]) -> dict[str, Any]:
    """Üst satır branş adları, alt satır Doğru/Yanlış/Boş/Net."""
    harita: dict[str, Any] = {"brans": {}, "toplam": {}}

    for idx, baslik in enumerate(alt):
        anahtar = normalize_ad(baslik)
        if not anahtar:
            continue
        if anahtar in {"ad soyad", "adsoyad", "isim", "ogrenci", "öğrenci"}:
            harita.setdefault("ad_soyad", idx)
        elif anahtar in {"sinif", "sınıf", "class"}:
            # İlk "Sınıf" kimlik kolonu; sıralama sütunundaki ikinciyi alma
            harita.setdefault("sinif", idx)
        elif anahtar in {"puan", "score"}:
            harita.setdefault("puan", idx)
        elif anahtar in {"toplam net", "genel net"}:
            harita["toplam"]["net"] = idx

    # Branş başlangıç kolonları (üst satırdaki dolu hücreler)
    brans_baslangic: list[tuple[int, str]] = []
    for idx, baslik in enumerate(ust):
        kod = _brans_kodu(baslik)
        if kod:
            brans_baslangic.append((idx, kod))
    brans_baslangic.sort(key=lambda x: x[0])

    def _kolon_bransi(col: int) -> str | None:
        aday = None
        for start, kod in brans_baslangic:
            if start <= col:
                aday = kod
            else:
                break
        # Puan / toplam net kolonlarından sonra branş yok
        puan_idx = harita.get("puan")
        toplam_net_idx = harita.get("toplam", {}).get("net")
        sinir = None
        for v in (puan_idx, toplam_net_idx):
            if v is not None and (sinir is None or v < sinir):
                sinir = v
        if sinir is not None and col >= sinir:
            return None
        return aday

    for idx, baslik in enumerate(alt):
        anahtar = normalize_ad(baslik)
        kod = _kolon_bransi(idx)
        if not kod:
            continue
        if anahtar in {"dogru", "doğru"}:
            harita["brans"].setdefault(kod, {})["dogru"] = idx
        elif anahtar in {"yanlis", "yanlış"}:
            harita["brans"].setdefault(kod, {})["yanlis"] = idx
        elif anahtar in {"bos", "boş"}:
            harita["brans"].setdefault(kod, {})["bos"] = idx
        elif anahtar == "net":
            harita["brans"].setdefault(kod, {})["net"] = idx

    return harita


@dataclass
class DenemeImportSatir:
    satir_no: int
    excel_ad_soyad: str
    sinif: str
    talebe_id: int | None = None
    eslesme: str = "yok"  # otomatik | alias | manuel | oneri | yok | atla
    branslar: dict = field(default_factory=dict)
    toplam: dict = field(default_factory=dict)
    puan: str = "0"
    hatalar: list[str] = field(default_factory=list)
    oneri_talebe_id: int | None = None
    oneri_ad_soyad: str = ""
    oneri_sinif: str = ""
    oneri_oran: int = 0
    oneriler: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "satir_no": self.satir_no,
            "excel_ad_soyad": self.excel_ad_soyad,
            "sinif": self.sinif,
            "talebe_id": self.talebe_id,
            "eslesme": self.eslesme,
            "branslar": self.branslar,
            "toplam": self.toplam,
            "puan": self.puan,
            "hatalar": self.hatalar,
            "oneri_talebe_id": self.oneri_talebe_id,
            "oneri_ad_soyad": self.oneri_ad_soyad,
            "oneri_sinif": self.oneri_sinif,
            "oneri_oran": self.oneri_oran,
            "oneriler": self.oneriler,
        }

    @classmethod
    def from_dict(cls, data: dict) -> DenemeImportSatir:
        alanlar = {
            "satir_no",
            "excel_ad_soyad",
            "sinif",
            "talebe_id",
            "eslesme",
            "branslar",
            "toplam",
            "puan",
            "hatalar",
            "oneri_talebe_id",
            "oneri_ad_soyad",
            "oneri_sinif",
            "oneri_oran",
            "oneriler",
        }
        return cls(**{k: data[k] for k in alanlar if k in data})


@dataclass
class DenemeImportOnizleme:
    satirlar: list[DenemeImportSatir] = field(default_factory=list)
    hatalar: list[str] = field(default_factory=list)
    format: str = ""

    @property
    def toplam_ogrenci(self) -> int:
        return len(self.satirlar)

    @property
    def eslesen(self) -> int:
        return sum(1 for s in self.satirlar if s.talebe_id)

    @property
    def eslesmeyen(self) -> int:
        return sum(
            1
            for s in self.satirlar
            if not s.talebe_id and s.eslesme not in {"atla"}
        )

    @property
    def oneri_bekleyen(self) -> int:
        return sum(1 for s in self.satirlar if s.eslesme == "oneri" and not s.talebe_id)

    def to_session(self) -> dict:
        return {
            "satirlar": [s.to_dict() for s in self.satirlar],
            "hatalar": self.hatalar,
            "format": self.format,
        }

    @classmethod
    def from_session(cls, data: dict) -> DenemeImportOnizleme:
        return cls(
            satirlar=[DenemeImportSatir.from_dict(s) for s in data.get("satirlar", [])],
            hatalar=data.get("hatalar", []),
            format=data.get("format", ""),
        )


def _satir_deger(satir: list[str], index: int | None) -> str:
    if index is None or index >= len(satir):
        return ""
    return satir[index]


def talebe_eslestir(excel_ad: str, sinif: str = "") -> tuple[Talebe | None, str, list[dict]]:
    """
    Dönüş: (talebe, eslesme_tipi, oneriler)
    eslesme_tipi: otomatik | alias | oneri | yok
    """
    norm = normalize_ad(excel_ad)
    if not norm:
        return None, "yok", []

    alias = (
        DenemeEslestirmeAlias.objects.filter(excel_adi=norm)
        .select_related("talebe")
        .first()
    )
    if alias and alias.talebe.aktif:
        return alias.talebe, "alias", []

    # Tam eşleşme (büyük/küçük harf)
    adaylar = list(Talebe.objects.filter(aktif=True, ad_soyad__iexact=excel_ad.strip()))
    if len(adaylar) == 1:
        return adaylar[0], "otomatik", []
    if len(adaylar) > 1 and sinif:
        sinif_norm = normalize_ad(sinif)
        filtre = [
            t
            for t in adaylar
            if sinif_norm
            in normalize_ad(str(t.sinif_sube.sinif if t.sinif_sube_id else t.sinif))
        ]
        if len(filtre) == 1:
            return filtre[0], "otomatik", []

    # Normalize tam eşleşme (aksan / boşluk)
    norm_adaylar = [
        t
        for t in Talebe.objects.filter(aktif=True).select_related("sinif_sube")
        if normalize_ad(t.ad_soyad or "") == norm
    ]
    if len(norm_adaylar) == 1:
        return norm_adaylar[0], "otomatik", []
    if len(norm_adaylar) > 1 and sinif:
        sinif_norm = normalize_ad(sinif)
        filtre = [
            t
            for t in norm_adaylar
            if sinif_norm
            in normalize_ad(str(t.sinif_sube.sinif if t.sinif_sube_id else t.sinif))
        ]
        if len(filtre) == 1:
            return filtre[0], "otomatik", []

    oneriler = talebe_benzer_adaylar(excel_ad)
    if oneriler:
        return None, "oneri", oneriler
    return None, "yok", []


def _satirdan_sonuclari_cek(
    satir: list[str], harita: dict[str, Any], kayit: DenemeImportSatir
) -> None:
    for kod, _, _ in BRANS_TANIMLARI:
        bh = harita.get("brans", {}).get(kod, {})
        if not bh:
            continue
        dogru = _tam_sayi(_satir_deger(satir, bh.get("dogru")))
        yanlis = _tam_sayi(_satir_deger(satir, bh.get("yanlis")))
        bos = _tam_sayi(_satir_deger(satir, bh.get("bos")))
        net_raw = _satir_deger(satir, bh.get("net"))
        net = _ondalik(net_raw) if net_raw else DenemeBransSonucu.net_hesapla(dogru, yanlis)
        if dogru or yanlis or bos or net_raw:
            kayit.branslar[kod] = {
                "dogru": dogru,
                "yanlis": yanlis,
                "bos": bos,
                "net": str(net),
            }

    th = harita.get("toplam", {})
    t_dogru = _tam_sayi(_satir_deger(satir, th.get("dogru")))
    t_yanlis = _tam_sayi(_satir_deger(satir, th.get("yanlis")))
    t_bos = _tam_sayi(_satir_deger(satir, th.get("bos")))
    t_net_raw = _satir_deger(satir, th.get("net"))
    if t_dogru or t_yanlis or t_bos or t_net_raw or kayit.branslar:
        if t_net_raw:
            t_net = _ondalik(t_net_raw)
        else:
            t_net = sum(
                (_ondalik(v.get("net", "0")) for v in kayit.branslar.values()),
                Decimal("0"),
            )
        if not (t_dogru or t_yanlis or t_bos) and kayit.branslar:
            t_dogru = sum(v.get("dogru", 0) for v in kayit.branslar.values())
            t_yanlis = sum(v.get("yanlis", 0) for v in kayit.branslar.values())
            t_bos = sum(v.get("bos", 0) for v in kayit.branslar.values())
        kayit.toplam = {
            "dogru": t_dogru,
            "yanlis": t_yanlis,
            "bos": t_bos,
            "net": str(t_net.quantize(Decimal("0.01"))),
        }

    kayit.puan = _satir_deger(satir, harita.get("puan")) or "0"


def deneme_excel_onizle(dosya) -> DenemeImportOnizleme:
    onizleme = DenemeImportOnizleme()
    try:
        satirlar = _excel_satirlari(dosya)
    except Exception as exc:
        onizleme.hatalar.append(f"Dosya okunamadı: {exc}")
        return onizleme

    if not satirlar:
        onizleme.hatalar.append("Excel dosyası boş.")
        return onizleme

    if _okyanus_format_mi(satirlar):
        harita = _okyanus_baslik_haritasi(satirlar[0], satirlar[1])
        veri_satirlari = satirlar[2:]
        baslangic_no = 3
        onizleme.format = "okyanus"
    else:
        harita = _duz_baslik_haritasi(satirlar[0])
        veri_satirlari = satirlar[1:]
        baslangic_no = 2
        onizleme.format = "duz"

    if "ad_soyad" not in harita:
        onizleme.hatalar.append("Ad Soyad sütunu bulunamadı.")
        return onizleme

    if not harita.get("brans"):
        onizleme.hatalar.append(
            "Ders (branş) sütunları bulunamadı. Okyanus genel sonuç raporu veya "
            "Türkçe Doğru / Yanlış / Boş / Net formatını kullanın."
        )
        return onizleme

    for i, satir in enumerate(veri_satirlari):
        satir_no = baslangic_no + i
        ad = _satir_deger(satir, harita.get("ad_soyad"))
        if not ad:
            continue
        sinif = _satir_deger(satir, harita.get("sinif"))
        kayit = DenemeImportSatir(satir_no=satir_no, excel_ad_soyad=ad, sinif=sinif)
        _satirdan_sonuclari_cek(satir, harita, kayit)

        talebe, eslesme, oneriler = talebe_eslestir(ad, sinif)
        kayit.eslesme = eslesme
        kayit.oneriler = oneriler
        if talebe:
            kayit.talebe_id = talebe.id
        elif oneriler:
            en_iyi = oneriler[0]
            kayit.oneri_talebe_id = en_iyi["id"]
            kayit.oneri_ad_soyad = en_iyi["ad_soyad"]
            kayit.oneri_sinif = en_iyi.get("sinif", "")
            kayit.oneri_oran = int(en_iyi.get("oran", 0))
            kayit.eslesme = "oneri"
            kayit.hatalar.append(
                f"«{ad}» sitedeki «{en_iyi['ad_soyad']}» (%{en_iyi['oran']}) "
                f"talebesine benziyor — doğru mu?"
            )

        onizleme.satirlar.append(kayit)

    if not onizleme.satirlar:
        onizleme.hatalar.append("Öğrenci satırı bulunamadı.")

    return onizleme


def session_key(deneme_id: int) -> str:
    return f"deneme_import_{deneme_id}"


@transaction.atomic
def deneme_sonuclari_aktar(
    deneme: DenemeSinavi,
    onizleme: DenemeImportOnizleme,
    user: User,
) -> tuple[int, list[str]]:
    hatalar: list[str] = []
    if deneme.durum == DenemeSinavi.Durum.AKTIF:
        return 0, ["Bu deneme zaten aktarılmış. Sonuçlar arşivlenir, tekrar yüklenemez."]

    oneri_bekleyen = [
        s for s in onizleme.satirlar if s.eslesme == "oneri" and not s.talebe_id
    ]
    if oneri_bekleyen:
        return 0, [
            f"{len(oneri_bekleyen)} satırda benzer isim önerisi onay bekliyor. "
            "«Evet, bu» veya manuel seçim yapın; misafirleri «Atla» diyebilirsiniz."
        ]

    eslesen = [s for s in onizleme.satirlar if s.talebe_id]
    if not eslesen:
        return 0, ["Aktarılacak eşleşmiş öğrenci yok."]

    atlanan = sum(1 for s in onizleme.satirlar if not s.talebe_id)

    DenemeSonucu.objects.filter(deneme=deneme).delete()
    kayit_sayisi = 0

    for satir in eslesen:
        talebe = Talebe.objects.get(pk=satir.talebe_id)
        toplam = satir.toplam or {}
        t_net = _ondalik(str(toplam.get("net", "0")))
        if not toplam and satir.branslar:
            t_net = sum(
                (_ondalik(v.get("net", "0")) for v in satir.branslar.values()),
                Decimal("0"),
            ).quantize(Decimal("0.01"))
            toplam = {
                "dogru": sum(v.get("dogru", 0) for v in satir.branslar.values()),
                "yanlis": sum(v.get("yanlis", 0) for v in satir.branslar.values()),
                "bos": sum(v.get("bos", 0) for v in satir.branslar.values()),
                "net": str(t_net),
            }

        sonuc = DenemeSonucu.objects.create(
            deneme=deneme,
            talebe=talebe,
            toplam_dogru=int(toplam.get("dogru", 0)),
            toplam_yanlis=int(toplam.get("yanlis", 0)),
            toplam_bos=int(toplam.get("bos", 0)),
            toplam_net=t_net,
            puan=_ondalik(satir.puan),
        )

        for kod, veri in satir.branslar.items():
            DenemeBransSonucu.objects.create(
                sonuc=sonuc,
                brans=kod,
                dogru=int(veri.get("dogru", 0)),
                yanlis=int(veri.get("yanlis", 0)),
                bos=int(veri.get("bos", 0)),
                net=_ondalik(str(veri.get("net", "0"))),
            )

        from takip.soru_takip_service import deneme_sonucu_soru_takibe_yansit

        deneme_sonucu_soru_takibe_yansit(user=user, deneme=deneme, sonuc=sonuc)

        norm = normalize_ad(satir.excel_ad_soyad)
        DenemeEslestirmeAlias.objects.update_or_create(
            excel_adi=norm,
            defaults={"talebe": talebe},
        )
        kayit_sayisi += 1

    deneme.durum = DenemeSinavi.Durum.AKTIF
    deneme.yukleyen = user
    deneme.yuklenme_zamani = timezone.now()
    deneme.save(
        update_fields=["durum", "yukleyen", "yuklenme_zamani", "guncellenme"]
    )
    if atlanan:
        hatalar.append(
            f"{atlanan} satır eşleşmediği / atlandığı için aktarılmadı."
        )
    return kayit_sayisi, hatalar
