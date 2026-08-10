"""
Ortak Excel rapor düzeni — örnek tablo stili.

- Üstte kurum logosu (ortalanmış, isim yazılmaz)
- Kalın başlık satırı + altın çizgi
- İnce yatay çizgiler, dikey kenarlık yok
- Durum kolonları hafif renkli; vurgu kolonları koyu kalın
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Sequence

from django.conf import settings
from django.http import HttpResponse


# Renkler (örnek düzene yakın)
_RENK_ALTIN = "B8956C"
_RENK_CIZGI = "D9DEE7"
_RENK_METIN = "1A1A1A"
_RENK_IKINCIL = "6B7280"
_RENK_DURUM_BG = "FCE8E8"
_RENK_DURUM_FG = "C62828"


@dataclass
class ExcelKolon:
    baslik: str
    genislik: float = 16
    tip: str = "metin"  # metin | durum | vurgu | sayi | ortala
    buyuk_harf: bool = False


@dataclass
class ExcelRapor:
    baslik: str
    kolonlar: list[ExcelKolon]
    satirlar: list[Sequence[Any]] = field(default_factory=list)
    alt_baslik: str = ""
    sayfa_adi: str = "Rapor"
    kurum: str = ""
    satir_yukseklik: float = 20
    metin_kaydir: bool = False
    kilitli: bool = True
    basliklari_buyuk_harf: bool = True


@dataclass
class ExcelSayfa:
    """Çok sayfalı çalışma kitabı için tek sayfa."""

    adi: str
    baslik: str
    kolonlar: list[ExcelKolon]
    satirlar: list[Sequence[Any]] = field(default_factory=list)
    alt_baslik: str = ""
    satir_yukseklik: float = 28
    metin_kaydir: bool = True
    kilitli: bool = True
    basliklari_buyuk_harf: bool = True


def _logo_yolu() -> Path | None:
    """Tam logo (Öğrenci Yurdu yazılı) tercih edilir; ikon yedek."""
    images = Path(settings.BASE_DIR) / "static" / "images"
    adaylar = [
        images / "cinili-saray-logo.jpeg",
        images / "cinili-saray-logo.jpg",
        images / "cinili-saray-logo.png",
        images / "cinili-saray-logo-icon.png",
        images / "cinili-saray-logo-icon.jpeg",
    ]
    for p in adaylar:
        if p.is_file():
            return p
    return None


def _sutun_px(genislik: float) -> int:
    """Excel sütun genişliğini yaklaşık piksele çevirir."""
    w = max(0.0, float(genislik))
    if w < 1:
        return max(1, int(w * 12))
    return max(1, int(w * 7 + 5))


def _logo_ortala_konum(
    kolonlar: list[ExcelKolon], hedef_w: int
) -> tuple[int, int, int]:
    """
    Ortalanmış logo için (başlangıç_sütun_0, colOff_px, toplam_px).
    Not: Excel'de colOff, yalnızca o sütunun genişliğini aşamaz;
    bu yüzden soldan büyük offset A1'e yazılmaz — doğru sütundan başlanmalı.
    """
    genislik_px = [_sutun_px(k.genislik) for k in kolonlar]
    toplam = sum(genislik_px) or hedef_w
    sol = max(0, (toplam - hedef_w) // 2)
    biriken = 0
    for i, w in enumerate(genislik_px):
        if biriken + w > sol:
            return i, sol - biriken, toplam
        biriken += w
    return max(0, len(kolonlar) - 1), 0, toplam


def _logo_ekle(ws, *, satir: int, kolon_sayisi: int, kolonlar: list[ExcelKolon]) -> tuple[bool, int]:
    """
    Logoyu verilen satırdaki hücreye bağlayıp yatay ortalar.
    Döner: (başarılı mı, logo yüksekliği px).
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    logo = _logo_yolu()
    if logo is None:
        return False, 0

    try:
        img = XLImage(str(logo))
    except Exception:
        return False, 0

    ham_w = float(img.width or 1024)
    ham_h = float(img.height or 640)
    hedef_w = 240
    hedef_h = max(48, int(hedef_w * ham_h / ham_w))
    img.width = hedef_w
    img.height = hedef_h

    start_col, col_off, _toplam = _logo_ortala_konum(kolonlar, hedef_w)
    # colOff o sütun genişliğini aşmasın
    max_off = max(0, _sutun_px(kolonlar[start_col].genislik) - 1)
    col_off = max(0, min(col_off, max_off))

    try:
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=start_col,
                colOff=pixels_to_EMU(col_off),
                row=max(0, satir - 1),
                rowOff=pixels_to_EMU(6),
            ),
            ext=XDRPositiveSize2D(
                pixels_to_EMU(hedef_w),
                pixels_to_EMU(hedef_h),
            ),
        )
        ws.add_image(img)
        return True, hedef_h
    except Exception:
        try:
            from openpyxl.utils import get_column_letter

            # Ortadaki sütuna bağla
            orta = get_column_letter(start_col + 1)
            ws.add_image(img, f"{orta}{satir}")
            return True, hedef_h
        except Exception:
            return False, 0


def _sayi_mi(deger: Any) -> Any:
    """Excel yeşil üçgenini azaltmak için sayıları sayı olarak yaz."""
    if deger is None or deger == "":
        return deger
    if isinstance(deger, bool):
        return deger
    if isinstance(deger, (int, float)):
        return deger
    try:
        from decimal import Decimal

        if isinstance(deger, Decimal):
            return float(deger)
    except Exception:
        pass
    if isinstance(deger, str):
        s = deger.strip().replace(",", ".")
        if not s:
            return deger
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            return deger
    return deger


def _sayfa_yaz(ws, rapor: ExcelRapor | ExcelSayfa) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    kolon_sayisi = max(1, len(rapor.kolonlar))

    for idx, kolon in enumerate(rapor.kolonlar, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = kolon.genislik

    logo_satir = 1
    ws.merge_cells(
        start_row=logo_satir,
        start_column=1,
        end_row=logo_satir,
        end_column=kolon_sayisi,
    )
    ws.cell(row=logo_satir, column=1, value="").alignment = Alignment(
        horizontal="center", vertical="center"
    )
    logo_ok, logo_h = _logo_ekle(
        ws, satir=logo_satir, kolon_sayisi=kolon_sayisi, kolonlar=rapor.kolonlar
    )
    ws.row_dimensions[logo_satir].height = (
        max(72, int(logo_h * 0.78) + 14) if logo_ok else 22
    )

    ws.row_dimensions[2].height = 10

    baslik_satir = 3
    ws.merge_cells(
        start_row=baslik_satir,
        start_column=1,
        end_row=baslik_satir,
        end_column=kolon_sayisi,
    )
    alt = rapor.baslik
    if rapor.alt_baslik:
        alt = f"{rapor.baslik}  ·  {rapor.alt_baslik}"
    alt_hucre = ws.cell(row=baslik_satir, column=1, value=alt)
    alt_hucre.font = Font(name="Calibri", size=11, color=_RENK_IKINCIL)
    alt_hucre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[baslik_satir].height = 20

    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 14

    header_row = 6
    ince = Side(style="thin", color=_RENK_CIZGI)
    ust_cizgi = Side(style="thin", color="C5CCD6")
    altin = Side(style="medium", color=_RENK_ALTIN)
    bos_kenar = Side(style=None)

    baslik_buyuk = getattr(rapor, "basliklari_buyuk_harf", True)
    for idx, kolon in enumerate(rapor.kolonlar, start=1):
        baslik_metin = str(kolon.baslik or "")
        if baslik_buyuk:
            baslik_metin = baslik_metin.upper()
        h = ws.cell(
            row=header_row,
            column=idx,
            value=baslik_metin,
        )
        h.font = Font(name="Calibri", bold=True, size=10, color=_RENK_METIN)
        h.alignment = Alignment(
            horizontal="center" if kolon.tip != "metin" else "left",
            vertical="center",
            wrap_text=True,
        )
        h.border = Border(
            left=bos_kenar,
            right=bos_kenar,
            top=ust_cizgi,
            bottom=altin,
        )

    ws.row_dimensions[header_row].height = 22

    durum_fill = PatternFill("solid", fgColor=_RENK_DURUM_BG)
    durum_font = Font(name="Calibri", size=10, color=_RENK_DURUM_FG, bold=True)
    metin_font = Font(name="Calibri", size=10, color=_RENK_METIN)
    ikincil_font = Font(name="Calibri", size=10, color=_RENK_IKINCIL)
    vurgu_font = Font(name="Calibri", size=11, color=_RENK_METIN, bold=True)
    kaydir = bool(getattr(rapor, "metin_kaydir", False))
    yukseklik = float(getattr(rapor, "satir_yukseklik", 20) or 20)

    for r_idx, satir in enumerate(rapor.satirlar, start=header_row + 1):
        degerler = list(satir)
        for c_idx, kolon in enumerate(rapor.kolonlar, start=1):
            ham = degerler[c_idx - 1] if c_idx - 1 < len(degerler) else ""
            if ham is None:
                ham = ""
            if kolon.buyuk_harf and isinstance(ham, str):
                ham = ham.upper()
            if kolon.tip in {"vurgu", "sayi", "ortala"}:
                ham = _sayi_mi(ham)

            hucre = ws.cell(row=r_idx, column=c_idx, value=ham)
            hucre.border = Border(
                left=bos_kenar,
                right=bos_kenar,
                top=bos_kenar,
                bottom=ince,
            )
            hucre.alignment = Alignment(
                horizontal="left" if kolon.tip == "metin" and c_idx == 1 else "center",
                vertical="center",
                wrap_text=kaydir,
            )

            if kolon.tip == "durum":
                hucre.fill = durum_fill
                hucre.font = durum_font
            elif kolon.tip in {"vurgu", "sayi"}:
                hucre.font = vurgu_font
                if isinstance(ham, float):
                    hucre.number_format = "0.00"
            elif kolon.tip == "ortala":
                hucre.font = ikincil_font
                if isinstance(ham, float):
                    hucre.number_format = "0.00"
            else:
                hucre.font = metin_font if c_idx == 1 else ikincil_font

        ws.row_dimensions[r_idx].height = yukseklik

    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_setup.orientation = "landscape" if kolon_sayisi > 6 else "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    if getattr(rapor, "kilitli", True):
        ws.protection.sheet = True
        ws.protection.objects = True
        ws.protection.scenarios = True
        ws.protection.enable()


def rapor_workbook_olustur(rapor: ExcelRapor):
    """ExcelRapor'dan Workbook üretir (doğrulama vb. eklemek için)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = (rapor.sayfa_adi or "Rapor")[:31]
    _sayfa_yaz(ws, rapor)
    return wb


def rapor_xlsx_olustur(rapor: ExcelRapor) -> bytes:
    wb = rapor_workbook_olustur(rapor)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# Başlık satırı (logo + boşluklardan sonra)
EXCEL_RAPOR_BASLIK_SATIRI = 6


def coklu_rapor_xlsx(sayfalar: Sequence[ExcelSayfa]) -> bytes:
    """Birden fazla sayfayı aynı logo / altın çizgi diliyle üretir."""
    from openpyxl import Workbook

    if not sayfalar:
        return rapor_xlsx_olustur(
            ExcelRapor(baslik="Rapor", kolonlar=[ExcelKolon("—")], satirlar=[])
        )

    wb = Workbook()
    ilk = True
    for sayfa in sayfalar:
        if ilk:
            ws = wb.active
            ilk = False
        else:
            ws = wb.create_sheet()
        ws.title = (sayfa.adi or "Sayfa")[:31]
        _sayfa_yaz(ws, sayfa)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_http_yanit(icerik: bytes, dosya_adi: str) -> HttpResponse:
    if not dosya_adi.lower().endswith(".xlsx"):
        dosya_adi = f"{dosya_adi}.xlsx"
    response = HttpResponse(
        icerik,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{dosya_adi}"'
    return response


def basit_rapor_xlsx(
    *,
    baslik: str,
    kolon_basliklari: Sequence[str],
    satirlar: Iterable[Sequence[Any]],
    alt_baslik: str = "",
    sayfa_adi: str = "Rapor",
    durum_kolonlari: Sequence[int] | None = None,
    vurgu_kolonlari: Sequence[int] | None = None,
    ortala_kolonlari: Sequence[int] | None = None,
    genislikler: Sequence[float] | None = None,
    buyuk_harf_kolonlari: Sequence[int] | None = None,
) -> bytes:
    """Kolon indeksleri 0 tabanlı."""
    durum_set = set(durum_kolonlari or [])
    vurgu_set = set(vurgu_kolonlari or [])
    ortala_set = set(ortala_kolonlari or [])
    buyuk_set = set(buyuk_harf_kolonlari or [])
    gen = list(genislikler) if genislikler else []

    kolonlar: list[ExcelKolon] = []
    for i, ad in enumerate(kolon_basliklari):
        if i in durum_set:
            tip = "durum"
        elif i in vurgu_set:
            tip = "vurgu"
        elif i in ortala_set:
            tip = "ortala"
        else:
            tip = "metin"
        kolonlar.append(
            ExcelKolon(
                baslik=ad,
                genislik=gen[i] if i < len(gen) else (22 if i == 0 else 14),
                tip=tip,
                buyuk_harf=i in buyuk_set,
            )
        )

    return rapor_xlsx_olustur(
        ExcelRapor(
            baslik=baslik,
            alt_baslik=alt_baslik,
            kolonlar=kolonlar,
            satirlar=[list(s) for s in satirlar],
            sayfa_adi=sayfa_adi,
        )
    )
